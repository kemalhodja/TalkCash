import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.i18n import t


class ExcelExportService:
    def generate_report(
        self,
        user_name: str,
        wallets: list[dict],
        transactions: list[dict],
        agenda_items: list[dict],
        locale: str = "tr",
    ) -> bytes:
        wb = Workbook()
        now = datetime.now().strftime("%d.%m.%Y %H:%M")

        ws_summary = wb.active
        ws_summary.title = t("export.sheet.summary", locale)
        ws_summary["A1"] = t("export.title", locale, name=user_name)
        ws_summary["A1"].font = Font(bold=True, size=16)
        ws_summary["A2"] = t("export.user", locale, name=user_name)
        ws_summary["A3"] = t("export.date", locale, date=now)

        ws_wallets = wb.create_sheet(t("export.sheet.wallets", locale))
        ws_wallets.append([
            t("export.col.wallet", locale),
            t("export.col.balance", locale),
            t("export.col.currency", locale),
            t("export.col.type", locale),
        ])
        for w in wallets:
            ws_wallets.append([w["name"], w["balance"], w.get("currency", "TRY"), w.get("type", "")])

        ws_tx = wb.create_sheet(t("export.sheet.transactions", locale))
        ws_tx.append([
            t("export.col.date", locale),
            t("export.col.category", locale),
            t("export.col.amount", locale),
            t("export.col.description", locale),
            t("export.col.place", locale),
        ])
        header_fill = PatternFill(start_color="00D4AA", end_color="00D4AA", fill_type="solid")
        for cell in ws_tx[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for tx in transactions:
            ws_tx.append([
                tx.get("date", ""), tx.get("category", ""),
                tx.get("amount", 0), tx.get("description", ""), tx.get("place", ""),
            ])

        ws_agenda = wb.create_sheet(t("export.sheet.agenda", locale))
        ws_agenda.append([
            t("export.col.title", locale),
            t("export.col.amount", locale),
            t("export.col.due_date", locale),
            t("export.col.status", locale),
        ])
        for a in agenda_items:
            ws_agenda.append([a["title"], a["amount"], a.get("due_date", ""), a.get("status", "")])

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
